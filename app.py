import streamlit as st
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO
import re

st.title("Extraction intelligente des tableaux PDF géotechnique")

def clean(v):
    if v is None:
        return ""
    return str(v).replace("\n", " ").strip()

def norm(txt):
    txt = clean(txt).lower()
    txt = txt.replace("è", "e").replace("é", "e").replace("ê", "e")
    txt = txt.replace("à", "a").replace("ç", "c")
    return txt

def detect_type(tableau):
    texte = " ".join(norm(c) for row in tableau[:4] for c in row if c)

    if "pressio" in texte or "pression limite" in texte or "module pressiometrique" in texte or " pl " in texte:
        return "Pressiometrique"

    if "compression" in texte or "resistance" in texte:
        return "Compression_simple"

    if "coord" in texte or (" x " in texte and " y " in texte):
        return "Coordonnees"

    if "lithologie" in texte or "formation" in texte or "nature" in texte:
        return "Lithologie"

    if "granulo" in texte or "tamis" in texte or "passant" in texte:
        return "Granulometrie"

    if "atterberg" in texte or "limite de liquidite" in texte or "ip" in texte:
        return "Atterberg"

    if "cbr" in texte or "ipi" in texte:
        return "CBR_IPI"

    return "Autre"

def get_header(tableau):
    for row in tableau:
        row_clean = [clean(c) for c in row]
        text = " ".join(row_clean).lower()
        if any(word in text for word in ["sondage", "profondeur", "pression", "module", "resistance", "coord", "formation"]):
            return row_clean
    return [f"Colonne_{i+1}" for i in range(max(len(r) for r in tableau))]

def same_header(h1, h2):
    a = " ".join(norm(x) for x in h1)
    b = " ".join(norm(x) for x in h2)
    return a == b or len(set(a.split()) & set(b.split())) >= 2

def clean_rows(tableau, header):
    rows = []
    for row in tableau:
        row_clean = [clean(c) for c in row]
        if not any(row_clean):
            continue
        if same_header(row_clean, header):
            continue
        rows.append(row_clean)
    return rows

def unique_sheet_name(name, existing):
    base = name[:25]
    final = base
    i = 1
    while final in existing:
        i += 1
        final = f"{base}_{i}"
    existing.add(final)
    return final

def style_sheet(ws):
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    fill = PatternFill("solid", fgColor="9DC3E6")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

pdf_file = st.file_uploader("Importer le PDF", type=["pdf"])

if pdf_file:
    groupes = {}

    with pdfplumber.open(pdf_file) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tableaux = page.extract_tables()

            for tableau in tableaux:
                if not tableau or len(tableau) < 2:
                    continue

                typ = detect_type(tableau)
                header = get_header(tableau)
                rows = clean_rows(tableau, header)

                if not rows:
                    continue

                cle_trouvee = None

                for cle, bloc in groupes.items():
                    if bloc["type"] == typ and same_header(bloc["header"], header):
                        cle_trouvee = cle
                        break

                if cle_trouvee is None:
                    cle_trouvee = f"{typ}_{len(groupes)+1}"
                    groupes[cle_trouvee] = {
                        "type": typ,
                        "header": header,
                        "rows": []
                    }

                groupes[cle_trouvee]["rows"].extend(rows)

    wb = Workbook()
    wb.remove(wb.active)

    existing_names = set()

    for cle, bloc in groupes.items():
        sheet_name = unique_sheet_name(bloc["type"], existing_names)
        ws = wb.create_sheet(sheet_name)

        ws.append(bloc["header"])

        for row in bloc["rows"]:
            ws.append(row)

        style_sheet(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.success(f"{len(groupes)} tableaux homogènes détectés et organisés.")

    st.download_button(
        label="Télécharger Excel organisé",
        data=output,
        file_name="extraction_geotech_intelligente.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
